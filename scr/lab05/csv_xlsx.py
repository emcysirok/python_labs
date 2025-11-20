import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyx1 ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
       # проверка
    if Path(csv_path).is_absolute() or Path(xlsx_path).is_absolute():
        raise ValueError("пути должн быть относительными")
    if not csv_path.endswith('.csv') or not xlsx_path.endswith('.xlsx'):
        raise ValueError("неверные расширения файлов")
    if not Path(csv_path).exists():
        raise FileNotFoundError(f"файл не найден: {csv_path}")
    
    # читаеем CSV
    with open(csv_path, "r", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    
    if not rows or not any(rows[0]):
        raise ValueError("пустой CSV или нет заголовка")
    
    # создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # записываем данные и считаем ширину
    max_lengths = []
    for row in rows:
        ws.append(row)
        for i, value in enumerate(row):
            if i >= len(max_lengths):
                max_lengths.append(0)
            max_lengths[i] = max(max_lengths[i], len(str(value or "")))
    
    # устанавливаем ширину колонок
    for i, length in enumerate(max_lengths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(length + 2, 8)
    
    # сохраняем
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)

if __name__ == "__main__":
    csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")